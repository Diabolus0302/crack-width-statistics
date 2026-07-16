# -*- coding: utf-8 -*-
from __future__ import annotations

import math
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageOps, ImageTk
from scipy import ndimage as ndi
from skimage.morphology import skeletonize

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except ImportError as exc:
    raise SystemExit("当前 Python 环境缺少 tkinter，请安装带 tkinter 的 Python。") from exc


APP_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
PROJECT_DIR = APP_DIR if getattr(sys, "frozen", False) else APP_DIR.parent
DEFAULT_DATA_DIR = PROJECT_DIR / "example-data-root" / "crop_20260219_194515"
IMAGE_SUFFIXES = {".tif", ".tiff", ".png", ".jpg", ".jpeg", ".bmp"}


@dataclass
class SegmentStat:
    segment_id: int
    network_id: int
    length_px: float
    width_px: float
    length_unit: float
    width_unit: float
    skeleton_points: int
    centroid_x: float
    centroid_y: float
    bbox_x: int
    bbox_y: int
    bbox_w: int
    bbox_h: int


@dataclass
class AnalysisResult:
    name: str
    mask: np.ndarray
    skeleton: np.ndarray
    widths_px: np.ndarray
    pixel_length_share: np.ndarray
    component_labels: np.ndarray
    segment_labels: np.ndarray
    branch_points: np.ndarray
    endpoint_points: np.ndarray
    isolated_points: np.ndarray
    segments: List[SegmentStat]
    bin_table: List[Dict[str, object]]
    width_summary: Dict[str, object]
    quality_metrics: Dict[str, object]
    total_length_px: float
    total_length_unit: float


def natural_key(text: str) -> List[object]:
    parts = re.split(r"(\d+)", text)
    return [int(p) if p.isdigit() else p.lower() for p in parts]


def read_scale_info(data_dir: Path) -> Tuple[float, float, str]:
    pixel_value = 100.0
    real_value = 1.0
    unit = "mm"
    info_path = data_dir / "scale_info.txt"
    if not info_path.exists():
        return pixel_value, real_value, unit

    text = info_path.read_text(encoding="utf-8", errors="ignore")
    unit_match = re.search(r"Unit:\s*(.+)", text)
    ppu_match = re.search(r"Pixels Per Unit:\s*([0-9.]+)", text)
    if unit_match:
        unit = unit_match.group(1).strip()
    if ppu_match:
        pixel_value = float(ppu_match.group(1))
        real_value = 1.0
    return pixel_value, real_value, unit


def list_images(folder: Path) -> List[Path]:
    if not folder.exists():
        return []
    return sorted(
        [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_SUFFIXES],
        key=lambda p: natural_key(p.name),
    )


def image_to_gray(path: Path) -> np.ndarray:
    with Image.open(path) as img:
        return np.asarray(ImageOps.grayscale(img), dtype=np.uint8)


def read_display_image(path: Optional[Path], fallback_shape: Tuple[int, int]) -> Image.Image:
    if path and path.exists():
        with Image.open(path) as img:
            return ImageOps.grayscale(img).convert("RGB")
    h, w = fallback_shape
    return Image.new("RGB", (w, h), "white")


def binary_mask_from_image(
    gray: np.ndarray,
    threshold: int,
    auto_foreground: bool,
    crack_is_dark: bool,
) -> np.ndarray:
    bright = gray >= threshold
    if auto_foreground:
        foreground = bright if bright.mean() <= 0.5 else ~bright
    else:
        foreground = ~bright if crack_is_dark else bright
    foreground = remove_small_components(foreground.astype(bool), max_size=2)
    return ndi.binary_fill_holes(foreground).astype(bool)


def remove_small_components(mask: np.ndarray, max_size: int) -> np.ndarray:
    labels, count = ndi.label(mask, structure=np.ones((3, 3), dtype=bool))
    if count == 0:
        return mask.astype(bool)
    sizes = np.bincount(labels.ravel())
    remove = sizes <= max_size
    remove[0] = False
    return mask & ~remove[labels]


def build_width_bins(range_min: float, range_max: float, step: float) -> List[Tuple[float, float]]:
    if range_max <= range_min:
        raise ValueError("宽度上限必须大于宽度下限。")
    if step <= 0:
        raise ValueError("区间步长必须大于 0。")

    bins: List[Tuple[float, float]] = []
    low = range_min
    guard = 0
    while low < range_max - 1e-12:
        high = min(range_max, low + step)
        bins.append((round(low, 10), round(high, 10)))
        low = high
        guard += 1
        if guard > 10000:
            raise ValueError("区间数量过多，请增大步长或缩小取值范围。")
    return bins


def unit_factor(pixel_value: float, real_value: float) -> float:
    if pixel_value <= 0:
        raise ValueError("像素值必须大于 0。")
    return real_value / pixel_value


def skeleton_edge_lengths(skeleton: np.ndarray) -> Tuple[float, np.ndarray]:
    coords = np.argwhere(skeleton)
    share = np.zeros(skeleton.shape, dtype=float)
    total = 0.0
    skel = skeleton.astype(bool)
    offsets = [(0, 1, 1.0), (1, 0, 1.0), (1, 1, math.sqrt(2.0)), (1, -1, math.sqrt(2.0))]
    h, w = skeleton.shape
    for y, x in coords:
        for dy, dx, weight in offsets:
            ny, nx = y + dy, x + dx
            if 0 <= ny < h and 0 <= nx < w and skel[ny, nx]:
                total += weight
                share[y, x] += weight / 2.0
                share[ny, nx] += weight / 2.0
    lonely = skel & (share == 0)
    share[lonely] = 1.0
    total += float(lonely.sum())
    return total, share


def skeleton_neighbor_count(skeleton: np.ndarray) -> np.ndarray:
    kernel = np.ones((3, 3), dtype=np.uint8)
    kernel[1, 1] = 0
    return ndi.convolve(skeleton.astype(np.uint8), kernel, mode="constant", cval=0)


def percent(part: float, total: float) -> float:
    if total <= 0:
        return 0.0
    return float(part / total * 100.0)


def weighted_percentile(values: np.ndarray, weights: np.ndarray, percentile_value: float) -> float:
    values = np.asarray(values, dtype=float)
    weights = np.asarray(weights, dtype=float)
    valid = np.isfinite(values) & np.isfinite(weights) & (weights > 0)
    if not np.any(valid):
        return 0.0
    values = values[valid]
    weights = weights[valid]
    order = np.argsort(values)
    values = values[order]
    weights = weights[order]
    cumulative = np.cumsum(weights)
    cutoff = cumulative[-1] * percentile_value / 100.0
    index = int(np.searchsorted(cumulative, cutoff, side="left"))
    return float(values[min(index, len(values) - 1)])


def weighted_width_summary(
    widths_px: np.ndarray,
    share: np.ndarray,
    skeleton: np.ndarray,
    px_to_unit: float,
    unit_name: str,
    exceed_threshold: float,
) -> Dict[str, object]:
    weights = share[skeleton]
    width_px_values = widths_px[skeleton]
    if weights.size == 0 or float(weights.sum()) <= 0:
        return {
            "总长度(px)": 0.0,
            f"总长度({unit_name})": 0.0,
            "长度加权平均宽度(px)": 0.0,
            f"长度加权平均宽度({unit_name})": 0.0,
            f"中位宽度({unit_name})": 0.0,
            f"P10宽度({unit_name})": 0.0,
            f"P90宽度({unit_name})": 0.0,
            f"最大宽度({unit_name})": 0.0,
            f"宽度标准差({unit_name})": 0.0,
            f"超阈值宽度阈值({unit_name})": exceed_threshold,
            "超阈值长度占比(%)": 0.0,
        }

    width_unit_values = width_px_values * px_to_unit
    total_length_px = float(weights.sum())
    total_length_unit = total_length_px * px_to_unit
    mean_px = float(np.average(width_px_values, weights=weights))
    mean_unit = mean_px * px_to_unit
    variance_unit = float(np.average((width_unit_values - mean_unit) ** 2, weights=weights))
    exceed_length_px = float(weights[width_unit_values >= exceed_threshold].sum())
    return {
        "总长度(px)": total_length_px,
        f"总长度({unit_name})": total_length_unit,
        "长度加权平均宽度(px)": mean_px,
        f"长度加权平均宽度({unit_name})": mean_unit,
        f"中位宽度({unit_name})": weighted_percentile(width_unit_values, weights, 50.0),
        f"P10宽度({unit_name})": weighted_percentile(width_unit_values, weights, 10.0),
        f"P90宽度({unit_name})": weighted_percentile(width_unit_values, weights, 90.0),
        f"最大宽度({unit_name})": float(width_unit_values.max()),
        f"宽度标准差({unit_name})": math.sqrt(max(variance_unit, 0.0)),
        f"超阈值宽度阈值({unit_name})": exceed_threshold,
        "超阈值长度占比(%)": percent(exceed_length_px, total_length_px),
    }


def build_bin_rows(
    skeleton: np.ndarray,
    widths_px: np.ndarray,
    share: np.ndarray,
    bins: List[Tuple[float, float]],
    px_to_unit: float,
    unit_name: str,
) -> Tuple[List[Dict[str, object]], float, float]:
    bin_rows: List[Dict[str, object]] = []
    width_units = widths_px * px_to_unit
    total_length_px = float(share[skeleton].sum())
    cumulative_px = 0.0
    range_min = bins[0][0] if bins else 0.0
    range_max = bins[-1][1] if bins else 0.0
    out_of_range = skeleton & ((width_units < range_min) | (width_units > range_max))
    out_of_range_length_px = float(share[out_of_range].sum())

    for idx, (low, high) in enumerate(bins):
        if idx == len(bins) - 1:
            in_bin = skeleton & (width_units >= low) & (width_units <= high)
        else:
            in_bin = skeleton & (width_units >= low) & (width_units < high)
        length_px = float(share[in_bin].sum())
        cumulative_px += length_px
        above_lower_px = float(share[skeleton & (width_units >= low)].sum())
        bin_rows.append(
            {
                "区间": f"{low:g}-{high:g}",
                f"区间下限({unit_name})": low,
                f"区间上限({unit_name})": high,
                "累计长度(px)": length_px,
                f"累计长度({unit_name})": length_px * px_to_unit,
                "长度占比(%)": percent(length_px, total_length_px),
                "累计占比(%)": percent(cumulative_px, total_length_px),
                f"≥区间下限长度({unit_name})": above_lower_px * px_to_unit,
                "≥区间下限占比(%)": percent(above_lower_px, total_length_px),
                "骨架点数": int(in_bin.sum()),
            }
        )
    return bin_rows, out_of_range_length_px, percent(out_of_range_length_px, total_length_px)


def main_width_bin_label(bin_rows: List[Dict[str, object]], unit_name: str) -> str:
    if not bin_rows:
        return "-"
    length_col = f"累计长度({unit_name})"
    row = max(bin_rows, key=lambda item: float(item.get(length_col, 0.0)))
    return f"{row['区间']} {unit_name}"


def format_value(value: object, digits: int = 4) -> str:
    if isinstance(value, float):
        if abs(value) >= 100:
            return f"{value:.2f}"
        return f"{value:.{digits}f}"
    return str(value)


def dict_to_rows(data: Dict[str, object]) -> List[Dict[str, object]]:
    return [{"指标": key, "数值": value} for key, value in data.items()]


def analyze_binary_image(
    image_path: Path,
    pixel_value: float,
    real_value: float,
    unit_name: str,
    bins: List[Tuple[float, float]],
    threshold: int,
    auto_foreground: bool,
    crack_is_dark: bool,
    min_skeleton_points: int,
    exceed_width_threshold: float,
) -> AnalysisResult:
    gray = image_to_gray(image_path)
    mask = binary_mask_from_image(gray, threshold, auto_foreground, crack_is_dark)
    raw_skeleton = skeletonize(mask)
    distances = ndi.distance_transform_edt(mask)
    widths_px = distances * 2.0
    _raw_total_length_px, share = skeleton_edge_lengths(raw_skeleton)
    px_to_unit = unit_factor(pixel_value, real_value)

    network_labels_raw, network_count = ndi.label(raw_skeleton, structure=np.ones((3, 3), dtype=bool))
    objects = ndi.find_objects(network_labels_raw)
    kept = np.zeros_like(raw_skeleton, dtype=bool)
    network_labels = np.zeros_like(network_labels_raw, dtype=np.int32)
    filtered_short_network_count = 0
    filtered_short_network_points = 0
    network_id = 1

    for label_id in range(1, network_count + 1):
        label_slice = objects[label_id - 1]
        if label_slice is None:
            continue
        component = network_labels_raw[label_slice] == label_id
        points = int(component.sum())
        if points < min_skeleton_points:
            filtered_short_network_count += 1
            filtered_short_network_points += points
            continue
        kept[label_slice] |= component
        sub_labels = network_labels[label_slice]
        sub_labels[component] = network_id
        network_labels[label_slice] = sub_labels
        network_id += 1

    skeleton = kept
    share = share * skeleton
    total_length_px = float(share.sum())
    total_length_unit = total_length_px * px_to_unit

    neighbors = skeleton_neighbor_count(skeleton)
    branch_points = skeleton & (neighbors >= 3)
    endpoint_points = skeleton & (neighbors == 1)
    isolated_points = skeleton & (neighbors == 0)

    split_seed = skeleton & ~branch_points
    split_labels_raw, split_count = ndi.label(split_seed, structure=np.ones((3, 3), dtype=bool))
    split_objects = ndi.find_objects(split_labels_raw)
    segment_labels = np.zeros_like(split_labels_raw, dtype=np.int32)
    segments: List[SegmentStat] = []
    segment_id = 1
    filtered_short_segment_count = 0
    filtered_short_segment_points = 0

    for label_id in range(1, split_count + 1):
        label_slice = split_objects[label_id - 1]
        if label_slice is None:
            continue
        component = split_labels_raw[label_slice] == label_id
        points = int(component.sum())
        if points < min_skeleton_points:
            filtered_short_segment_count += 1
            filtered_short_segment_points += points
            continue
        component_share = share[label_slice] * component
        length_px = float(component_share.sum())
        if length_px <= 0:
            continue
        width_px = float((widths_px[label_slice] * component_share).sum() / length_px)
        coords = np.argwhere(component)
        minr = label_slice[0].start
        minc = label_slice[1].start
        maxr = label_slice[0].stop
        maxc = label_slice[1].stop
        cy = float(coords[:, 0].mean() + minr)
        cx = float(coords[:, 1].mean() + minc)
        network_values = network_labels[label_slice][component]
        network_values = network_values[network_values > 0]
        parent_network_id = int(np.bincount(network_values).argmax()) if network_values.size else 0
        segments.append(
            SegmentStat(
                segment_id=segment_id,
                network_id=parent_network_id,
                length_px=length_px,
                width_px=width_px,
                length_unit=length_px * px_to_unit,
                width_unit=width_px * px_to_unit,
                skeleton_points=points,
                centroid_x=float(cx),
                centroid_y=float(cy),
                bbox_x=int(minc),
                bbox_y=int(minr),
                bbox_w=int(maxc - minc),
                bbox_h=int(maxr - minr),
            )
        )
        sub_labels = segment_labels[label_slice]
        sub_labels[component] = segment_id
        segment_labels[label_slice] = sub_labels
        segment_id += 1

    bin_rows, out_of_range_length_px, out_of_range_percent = build_bin_rows(
        skeleton=skeleton,
        widths_px=widths_px,
        share=share,
        bins=bins,
        px_to_unit=px_to_unit,
        unit_name=unit_name,
    )
    width_summary = weighted_width_summary(
        widths_px=widths_px,
        share=share,
        skeleton=skeleton,
        px_to_unit=px_to_unit,
        unit_name=unit_name,
        exceed_threshold=exceed_width_threshold,
    )
    segment_length_px = float(sum(seg.length_px for seg in segments))
    branch_length_px = max(0.0, total_length_px - segment_length_px)
    qc_flags = []
    if filtered_short_network_count:
        qc_flags.append("短网络被过滤")
    if filtered_short_segment_count:
        qc_flags.append("短切段被过滤")
    if branch_points.any():
        qc_flags.append("存在分叉")
    if isolated_points.any():
        qc_flags.append("存在孤立点")
    if out_of_range_percent > 5:
        qc_flags.append("宽度超出范围较多")
    if int(endpoint_points.sum()) > max(20, len(segments) * 3):
        qc_flags.append("端点数量偏多")
    quality_metrics = {
        "原始裂纹面积(px^2)": int(mask.sum()),
        "原始骨架点数": int(raw_skeleton.sum()),
        "保留骨架点数": int(skeleton.sum()),
        "合并裂纹网络数": max(0, network_id - 1),
        "分叉切段数": len(segments),
        "被过滤短网络数": filtered_short_network_count,
        "被过滤短网络点数": filtered_short_network_points,
        "被过滤短切段数": filtered_short_segment_count,
        "被过滤短切段点数": filtered_short_segment_points,
        "端点数": int(endpoint_points.sum()),
        "分叉点数": int(branch_points.sum()),
        "孤立点数": int(isolated_points.sum()),
        "超出宽度范围长度(px)": out_of_range_length_px,
        f"超出宽度范围长度({unit_name})": out_of_range_length_px * px_to_unit,
        "超出宽度范围占比(%)": out_of_range_percent,
        "分叉点未分配长度(px)": branch_length_px,
        f"分叉点未分配长度({unit_name})": branch_length_px * px_to_unit,
        "质检提示": "；".join(qc_flags) if qc_flags else "通过",
    }

    return AnalysisResult(
        name=image_path.stem,
        mask=mask,
        skeleton=skeleton,
        widths_px=widths_px,
        pixel_length_share=share,
        component_labels=network_labels,
        segment_labels=segment_labels,
        branch_points=branch_points,
        endpoint_points=endpoint_points,
        isolated_points=isolated_points,
        segments=segments,
        bin_table=bin_rows,
        width_summary=width_summary,
        quality_metrics=quality_metrics,
        total_length_px=total_length_px,
        total_length_unit=total_length_unit,
    )


def segment_columns(unit_name: str) -> List[str]:
    return [
        "裂纹段编号",
        "合并网络编号",
        "长度(px)",
        "宽度(px)",
        f"长度({unit_name})",
        f"宽度({unit_name})",
        "骨架点数",
        "质心X",
        "质心Y",
        "边界框X",
        "边界框Y",
        "边界框宽",
        "边界框高",
    ]


def segments_to_rows(segments: List[SegmentStat], unit_name: str) -> List[Dict[str, object]]:
    return [
        {
            "裂纹段编号": s.segment_id,
            "合并网络编号": s.network_id,
            "长度(px)": s.length_px,
            "宽度(px)": s.width_px,
            f"长度({unit_name})": s.length_unit,
            f"宽度({unit_name})": s.width_unit,
            "骨架点数": s.skeleton_points,
            "质心X": s.centroid_x,
            "质心Y": s.centroid_y,
            "边界框X": s.bbox_x,
            "边界框Y": s.bbox_y,
            "边界框宽": s.bbox_w,
            "边界框高": s.bbox_h,
        }
        for s in segments
    ]


def write_table(ws, rows: List[Dict[str, object]], start_row: int, columns: Optional[List[str]] = None) -> int:
    if columns is None:
        columns = list(rows[0].keys()) if rows else []
    if not columns:
        return start_row
    for col_idx, column in enumerate(columns, start=1):
        ws.cell(row=start_row, column=col_idx, value=column)
    for row_idx, row_data in enumerate(rows, start=start_row + 1):
        for col_idx, column in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(column))
    return start_row + len(rows) + 1


def unique_sheet_name(workbook: Workbook, name: str) -> str:
    base = CrackStatsApp.safe_sheet_name(name)
    candidate = base
    suffix = 1
    while candidate in workbook.sheetnames:
        suffix_text = f"_{suffix}"
        candidate = f"{base[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    return candidate


def run_self_test(data_dir: Path, output_dir: Path) -> int:
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / "packaged_self_test_report.json"
    workbook_path = output_dir / "packaged_self_test.xlsx"
    report: Dict[str, object] = {
        "status": "error",
        "data_dir": str(data_dir),
        "workbook": str(workbook_path),
    }
    try:
        pixel_value, real_value, unit_name = read_scale_info(data_dir)
        bins = build_width_bins(0.0, 3.0, 0.5)
        binary_dir = data_dir / "\u4e8c\u503c\u5316\u56fe"
        binary_images = list_images(binary_dir)
        if not binary_images:
            raise FileNotFoundError(f"No binary images found in {binary_dir}")

        result = analyze_binary_image(
            image_path=binary_images[0],
            pixel_value=pixel_value,
            real_value=real_value,
            unit_name=unit_name,
            bins=bins,
            threshold=128,
            auto_foreground=True,
            crack_is_dark=True,
            min_skeleton_points=5,
            exceed_width_threshold=1.5,
        )

        workbook = Workbook()
        workbook.remove(workbook.active)
        sheet = workbook.create_sheet(unique_sheet_name(workbook, result.name))
        next_row = write_table(sheet, segments_to_rows(result.segments, unit_name), 1, segment_columns(unit_name))
        next_row = write_table(sheet, result.bin_table, next_row + 2)
        next_row = write_table(sheet, dict_to_rows(result.width_summary), next_row + 2, ["指标", "数值"])
        write_table(sheet, dict_to_rows(result.quality_metrics), next_row + 2, ["指标", "数值"])
        workbook.save(workbook_path)

        mean_col = f"长度加权平均宽度({unit_name})"
        report.update(
            {
                "status": "ok",
                "image": str(binary_images[0]),
                "segment_count": len(result.segments),
                "total_length_px": result.total_length_px,
                "total_length_unit": result.total_length_unit,
                "bin_count": len(result.bin_table),
                "weighted_mean_width": result.width_summary.get(mean_col),
                "branch_points": result.quality_metrics.get("分叉点数"),
                "qc_note": result.quality_metrics.get("质检提示"),
            }
        )
        return 0
    except Exception as exc:
        report["error"] = repr(exc)
        return 1
    finally:
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def colors_for_bins(count: int) -> List[Tuple[int, int, int]]:
    base = [
        (61, 90, 254),
        (0, 172, 193),
        (67, 160, 71),
        (251, 192, 45),
        (245, 124, 0),
        (229, 57, 53),
        (142, 36, 170),
        (93, 64, 55),
    ]
    if count <= len(base):
        return base[:count]
    colors = []
    for i in range(count):
        hue = i / max(count, 1)
        r = int(127 + 100 * math.sin(2 * math.pi * hue))
        g = int(127 + 100 * math.sin(2 * math.pi * (hue + 0.33)))
        b = int(127 + 100 * math.sin(2 * math.pi * (hue + 0.66)))
        colors.append((r, g, b))
    return colors


def draw_skeleton_pixels(
    image: Image.Image,
    result: AnalysisResult,
    bins: List[Tuple[float, float]],
    pixel_value: float,
    real_value: float,
    mode: str,
    show_labels: bool,
    alpha: float = 1.0,
    line_width: int = 3,
    selected_segment_id: Optional[int] = None,
    selected_bin_index: Optional[int] = None,
) -> Image.Image:
    out = image.convert("RGB")
    arr = np.asarray(out).copy().astype(np.float32)
    px_to_unit = unit_factor(pixel_value, real_value)
    yx = np.argwhere(result.skeleton)
    alpha = min(1.0, max(0.05, float(alpha)))
    line_width = max(1, int(line_width))
    highlight_mask = None
    widths_unit = result.widths_px[result.skeleton] * px_to_unit

    if selected_segment_id is not None:
        highlight_mask = result.segment_labels[result.skeleton] == selected_segment_id
    elif selected_bin_index is not None and 0 <= selected_bin_index < len(bins):
        low, high = bins[selected_bin_index]
        if selected_bin_index == len(bins) - 1:
            highlight_mask = (widths_unit >= low) & (widths_unit <= high)
        else:
            highlight_mask = (widths_unit >= low) & (widths_unit < high)

    if mode == "blue":
        color_lookup = np.tile(np.array([[20, 35, 240]], dtype=np.float32), (len(yx), 1))
    else:
        bin_colors = np.array(colors_for_bins(len(bins)), dtype=np.float32)
        color_lookup = np.tile(np.array([[170, 170, 170]], dtype=np.float32), (len(yx), 1))
        for idx, (low, high) in enumerate(bins):
            if idx == len(bins) - 1:
                in_bin = (widths_unit >= low) & (widths_unit <= high)
            else:
                in_bin = (widths_unit >= low) & (widths_unit < high)
            color_lookup[in_bin] = bin_colors[idx]

    if highlight_mask is not None:
        color_lookup[:] = np.array([185, 185, 185], dtype=np.float32)
        color_lookup[highlight_mask] = np.array([255, 40, 20], dtype=np.float32)

    h, w = result.skeleton.shape
    radius = max(0, line_width // 2)
    for (y, x), color in zip(yx, color_lookup):
        y0, y1 = max(0, y - radius), min(h, y + radius + 1)
        x0, x1 = max(0, x - radius), min(w, x + radius + 1)
        arr[y0:y1, x0:x1] = arr[y0:y1, x0:x1] * (1.0 - alpha) + color * alpha

    out = Image.fromarray(np.clip(arr, 0, 255).astype(np.uint8))
    if show_labels:
        draw = ImageDraw.Draw(out)
        for seg in result.segments:
            fill = (255, 0, 0) if selected_segment_id in (None, seg.segment_id) else (120, 120, 120)
            draw.text((seg.centroid_x + 4, seg.centroid_y + 4), str(seg.segment_id), fill=fill)
    return out


def mask_to_display(mask: np.ndarray, white_background: bool) -> Image.Image:
    if white_background:
        arr = np.where(mask, 0, 255).astype(np.uint8)
    else:
        arr = np.where(mask, 255, 0).astype(np.uint8)
    return Image.fromarray(arr, mode="L").convert("RGB")


class ImagePanel(ttk.Frame):
    def __init__(self, master: tk.Misc, title: str) -> None:
        super().__init__(master)
        self.title_var = tk.StringVar(value=title)
        ttk.Label(self, textvariable=self.title_var, anchor="center").pack(fill="x", pady=(6, 4))
        self.canvas = tk.Canvas(self, bg="white", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        self._photo: Optional[ImageTk.PhotoImage] = None
        self._image: Optional[Image.Image] = None
        self.zoom = 1.0
        self.offset_x = 0.0
        self.offset_y = 0.0
        self._drag_start: Optional[Tuple[int, int, float, float]] = None
        self.canvas.bind("<Configure>", lambda _event: self.redraw())
        self.canvas.bind("<MouseWheel>", self.on_mousewheel)
        self.canvas.bind("<Button-4>", lambda event: self.zoom_at(event, 1.15))
        self.canvas.bind("<Button-5>", lambda event: self.zoom_at(event, 1 / 1.15))
        self.canvas.bind("<ButtonPress-1>", self.start_pan)
        self.canvas.bind("<B1-Motion>", self.pan)
        self.canvas.bind("<Double-Button-1>", lambda _event: self.fit_to_window())

    def set_image(self, image: Image.Image) -> None:
        self._image = image
        self.redraw()

    def fit_to_window(self) -> None:
        self.zoom = 1.0
        self.offset_x = 0.0
        self.offset_y = 0.0
        self.redraw()

    def on_mousewheel(self, event: tk.Event) -> None:
        factor = 1.15 if event.delta > 0 else 1 / 1.15
        self.zoom_at(event, factor)

    def zoom_at(self, event: tk.Event, factor: float) -> None:
        old_zoom = self.zoom
        self.zoom = min(12.0, max(0.1, self.zoom * factor))
        if self.zoom == old_zoom:
            return
        cx = max(self.canvas.winfo_width(), 10) / 2
        cy = max(self.canvas.winfo_height(), 10) / 2
        self.offset_x = (self.offset_x + cx - event.x) * (self.zoom / old_zoom) - (cx - event.x)
        self.offset_y = (self.offset_y + cy - event.y) * (self.zoom / old_zoom) - (cy - event.y)
        self.redraw()

    def start_pan(self, event: tk.Event) -> None:
        self._drag_start = (event.x, event.y, self.offset_x, self.offset_y)

    def pan(self, event: tk.Event) -> None:
        if self._drag_start is None:
            return
        start_x, start_y, old_offset_x, old_offset_y = self._drag_start
        self.offset_x = old_offset_x + event.x - start_x
        self.offset_y = old_offset_y + event.y - start_y
        self.redraw()

    def redraw(self) -> None:
        self.canvas.delete("all")
        if self._image is None:
            return
        cw = max(self.canvas.winfo_width(), 10)
        ch = max(self.canvas.winfo_height(), 10)
        iw, ih = self._image.size
        scale = min(cw / iw, ch / ih) * self.zoom
        nw, nh = max(1, int(iw * scale)), max(1, int(ih * scale))
        resized = self._image.resize((nw, nh), Image.Resampling.LANCZOS)
        self._photo = ImageTk.PhotoImage(resized)
        self.canvas.create_image(cw // 2 + self.offset_x, ch // 2 + self.offset_y, image=self._photo, anchor="center")


class CrackStatsApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("裂纹宽度分段统计")
        self.geometry("1580x860")
        self.minsize(1180, 680)

        self.data_dir = DEFAULT_DATA_DIR
        self.original_dir = self.data_dir / "原图"
        self.binary_dir = self.data_dir / "二值化图"
        self.binary_images: List[Path] = []
        self.original_map: Dict[str, Path] = {}
        self.index = 0
        self.current_result: Optional[AnalysisResult] = None
        self.result_cache: Dict[Tuple[str, str], AnalysisResult] = {}
        self.selected_segment_id: Optional[int] = None
        self.selected_bin_index: Optional[int] = None
        self._quality_preview_images: List[ImageTk.PhotoImage] = []

        pixel_value, real_value, unit_name = read_scale_info(self.data_dir)
        self.pixel_value_var = tk.StringVar(value=f"{pixel_value:g}")
        self.real_value_var = tk.StringVar(value=f"{real_value:g}")
        self.unit_var = tk.StringVar(value=unit_name)
        self.threshold_var = tk.StringVar(value="128")
        self.width_min_var = tk.StringVar(value="0")
        self.width_max_var = tk.StringVar(value="3")
        self.width_step_var = tk.StringVar(value="0.5")
        self.exceed_width_var = tk.StringVar(value="1.5")
        self.min_points_var = tk.StringVar(value="5")
        self.overlay_alpha_var = tk.DoubleVar(value=0.95)
        self.line_width_var = tk.DoubleVar(value=3)
        self.white_bg_var = tk.BooleanVar(value=True)
        self.show_labels_var = tk.BooleanVar(value=False)
        self.auto_foreground_var = tk.BooleanVar(value=True)
        self.crack_dark_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="")
        self.summary_var = tk.StringVar(value="")

        self._build_ui()
        self.reload_images()
        self.analyze_current()

    def _build_ui(self) -> None:
        root = ttk.Frame(self)
        root.pack(fill="both", expand=True, padx=12, pady=12)
        root.columnconfigure(0, weight=0, minsize=330)
        root.columnconfigure(1, weight=1)
        root.columnconfigure(2, weight=1)
        root.rowconfigure(0, weight=1)

        controls = ttk.Frame(root)
        controls.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        controls.columnconfigure(1, weight=1)

        ttk.Button(controls, text="选择数据目录", command=self.choose_data_dir).grid(row=0, column=0, columnspan=2, sticky="ew")
        self.path_label = ttk.Label(controls, text=str(self.data_dir), wraplength=310)
        self.path_label.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(4, 10))

        row = 2
        for label, var in [
            ("像素值(px)", self.pixel_value_var),
            ("对应实际尺度", self.real_value_var),
            ("单位名", self.unit_var),
            ("二值阈值", self.threshold_var),
            ("最小骨架点数", self.min_points_var),
            ("宽度下限", self.width_min_var),
            ("宽度上限", self.width_max_var),
            ("区间步长", self.width_step_var),
            ("超阈值宽度", self.exceed_width_var),
        ]:
            ttk.Label(controls, text=label).grid(row=row, column=0, sticky="w", pady=3)
            ttk.Entry(controls, textvariable=var).grid(row=row, column=1, sticky="ew", pady=3)
            row += 1

        ttk.Checkbutton(controls, text="二值视图白底", variable=self.white_bg_var, command=self.refresh_views).grid(
            row=row, column=0, columnspan=2, sticky="w", pady=(8, 2)
        )
        row += 1
        ttk.Checkbutton(controls, text="显示编号", variable=self.show_labels_var, command=self.refresh_views).grid(
            row=row, column=0, columnspan=2, sticky="w", pady=2
        )
        row += 1
        ttk.Checkbutton(controls, text="自动识别裂纹颜色", variable=self.auto_foreground_var, command=self.analyze_current).grid(
            row=row, column=0, columnspan=2, sticky="w", pady=2
        )
        row += 1
        ttk.Checkbutton(controls, text="手动模式下裂纹为深色", variable=self.crack_dark_var, command=self.analyze_current).grid(
            row=row, column=0, columnspan=2, sticky="w", pady=2
        )
        row += 1

        ttk.Label(controls, text="骨架透明度").grid(row=row, column=0, sticky="w", pady=(8, 2))
        ttk.Scale(controls, from_=0.1, to=1.0, variable=self.overlay_alpha_var, command=lambda _value: self.refresh_views()).grid(
            row=row, column=1, sticky="ew", pady=(8, 2)
        )
        row += 1
        ttk.Label(controls, text="骨架线宽").grid(row=row, column=0, sticky="w", pady=2)
        ttk.Scale(controls, from_=1, to=7, variable=self.line_width_var, command=lambda _value: self.refresh_views()).grid(
            row=row, column=1, sticky="ew", pady=2
        )
        row += 1

        ttk.Button(controls, text="重新统计当前图", command=self.analyze_current).grid(row=row, column=0, columnspan=2, sticky="ew", pady=(10, 4))
        row += 1
        ttk.Button(controls, text="批量质检预览", command=self.show_quality_preview).grid(row=row, column=0, columnspan=2, sticky="ew", pady=(0, 4))
        row += 1
        ttk.Button(controls, text="一键输出全部", command=self.export_all).grid(row=row, column=0, columnspan=2, sticky="ew")
        row += 1

        nav = ttk.Frame(controls)
        nav.grid(row=row, column=0, columnspan=2, sticky="ew", pady=(12, 4))
        nav.columnconfigure((0, 1), weight=1)
        ttk.Button(nav, text="← 上一张", command=self.prev_image).grid(row=0, column=0, sticky="ew", padx=(0, 4))
        ttk.Button(nav, text="下一张 →", command=self.next_image).grid(row=0, column=1, sticky="ew", padx=(4, 0))
        row += 1

        self.counter_var = tk.StringVar(value="")
        ttk.Label(controls, textvariable=self.counter_var, anchor="center").grid(row=row, column=0, columnspan=2, sticky="ew")
        row += 1

        ttk.Label(controls, textvariable=self.summary_var, wraplength=310, foreground="#222").grid(
            row=row, column=0, columnspan=2, sticky="ew", pady=(8, 0)
        )
        row += 1

        ttk.Label(controls, textvariable=self.status_var, wraplength=310, foreground="#555").grid(
            row=row, column=0, columnspan=2, sticky="ew", pady=(10, 0)
        )

        center = ttk.Frame(root)
        center.grid(row=0, column=1, sticky="nsew", padx=(0, 10))
        center.rowconfigure(0, weight=1)
        center.columnconfigure(0, weight=1)
        self.overlay_panel = ImagePanel(center, "叠加图")
        self.overlay_panel.grid(row=0, column=0, sticky="nsew")

        right = ttk.Frame(root)
        right.grid(row=0, column=2, sticky="nsew")
        right.rowconfigure(0, weight=3)
        right.rowconfigure(1, weight=0)
        right.rowconfigure(2, weight=2)
        right.columnconfigure(0, weight=1)
        self.binary_panel = ImagePanel(right, "二值视图（分段着色）")
        self.binary_panel.grid(row=0, column=0, sticky="nsew")

        self.legend_canvas = tk.Canvas(right, height=54, bg="#f4f4f4", highlightthickness=0)
        self.legend_canvas.grid(row=1, column=0, sticky="ew", pady=(4, 4))
        self.legend_canvas.bind("<Configure>", lambda _event: self.refresh_legend())

        table_frame = ttk.Frame(right)
        table_frame.grid(row=2, column=0, sticky="nsew", pady=(4, 0))
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        self.table_notebook = ttk.Notebook(table_frame)
        self.table_notebook.grid(row=0, column=0, sticky="nsew")

        segment_tab = ttk.Frame(self.table_notebook)
        segment_tab.rowconfigure(0, weight=1)
        segment_tab.columnconfigure(0, weight=1)
        self.table_notebook.add(segment_tab, text="裂纹切段")
        columns = ("id", "network", "length_px", "width_px", "length_unit", "width_unit", "points")
        self.segment_table = ttk.Treeview(segment_tab, columns=columns, show="headings", height=8)
        for col, text, width in [
            ("id", "段ID", 70),
            ("network", "网络ID", 70),
            ("length_px", "长度(px)", 95),
            ("width_px", "宽度(px)", 95),
            ("length_unit", "长度", 95),
            ("width_unit", "宽度", 95),
            ("points", "骨架点数", 85),
        ]:
            self.segment_table.heading(col, text=text)
            self.segment_table.column(col, width=width, anchor="center")
        self.segment_table.grid(row=0, column=0, sticky="nsew")
        segment_ybar = ttk.Scrollbar(segment_tab, orient="vertical", command=self.segment_table.yview)
        segment_ybar.grid(row=0, column=1, sticky="ns")
        self.segment_table.configure(yscrollcommand=segment_ybar.set)
        self.segment_table.bind("<<TreeviewSelect>>", self.on_segment_select)
        self.table = self.segment_table

        bin_tab = ttk.Frame(self.table_notebook)
        bin_tab.rowconfigure(0, weight=1)
        bin_tab.columnconfigure(0, weight=1)
        self.table_notebook.add(bin_tab, text="宽度区间")
        bin_columns = ("range", "length", "pct", "cum_pct", "above_pct", "points")
        self.bin_table_widget = ttk.Treeview(bin_tab, columns=bin_columns, show="headings", height=8)
        for col, text, width in [
            ("range", "区间", 95),
            ("length", "长度", 105),
            ("pct", "占比(%)", 80),
            ("cum_pct", "累计(%)", 80),
            ("above_pct", "≥下限(%)", 85),
            ("points", "骨架点数", 85),
        ]:
            self.bin_table_widget.heading(col, text=text)
            self.bin_table_widget.column(col, width=width, anchor="center")
        self.bin_table_widget.grid(row=0, column=0, sticky="nsew")
        bin_ybar = ttk.Scrollbar(bin_tab, orient="vertical", command=self.bin_table_widget.yview)
        bin_ybar.grid(row=0, column=1, sticky="ns")
        self.bin_table_widget.configure(yscrollcommand=bin_ybar.set)
        self.bin_table_widget.bind("<<TreeviewSelect>>", self.on_bin_select)

        qc_tab = ttk.Frame(self.table_notebook)
        qc_tab.rowconfigure(0, weight=1)
        qc_tab.columnconfigure(0, weight=1)
        self.table_notebook.add(qc_tab, text="质量指标")
        self.qc_table = ttk.Treeview(qc_tab, columns=("metric", "value"), show="headings", height=8)
        self.qc_table.heading("metric", text="指标")
        self.qc_table.heading("value", text="数值")
        self.qc_table.column("metric", width=210, anchor="w")
        self.qc_table.column("value", width=220, anchor="w")
        self.qc_table.grid(row=0, column=0, sticky="nsew")
        qc_ybar = ttk.Scrollbar(qc_tab, orient="vertical", command=self.qc_table.yview)
        qc_ybar.grid(row=0, column=1, sticky="ns")
        self.qc_table.configure(yscrollcommand=qc_ybar.set)

    def choose_data_dir(self) -> None:
        selected = filedialog.askdirectory(initialdir=str(self.data_dir), title="选择包含原图和二值化图的目录")
        if not selected:
            return
        self.data_dir = Path(selected)
        self.original_dir = self.data_dir / "原图"
        self.binary_dir = self.data_dir / "二值化图"
        pixel_value, real_value, unit_name = read_scale_info(self.data_dir)
        self.pixel_value_var.set(f"{pixel_value:g}")
        self.real_value_var.set(f"{real_value:g}")
        self.unit_var.set(unit_name)
        self.reload_images()
        self.analyze_current()

    def reload_images(self) -> None:
        self.binary_images = list_images(self.binary_dir)
        originals = list_images(self.original_dir)
        self.original_map = {p.name: p for p in originals}
        self.index = min(self.index, max(0, len(self.binary_images) - 1))
        self.path_label.configure(text=str(self.data_dir))
        self.result_cache.clear()
        self.update_counter()

    def update_counter(self) -> None:
        if not self.binary_images:
            self.counter_var.set("未找到二值化图")
            return
        self.counter_var.set(f"{self.binary_images[self.index].stem}  ({self.index + 1}/{len(self.binary_images)})")

    def current_parameters(self) -> Tuple[float, float, str, List[Tuple[float, float]], int, int, float]:
        pixel_value = float(self.pixel_value_var.get())
        real_value = float(self.real_value_var.get())
        unit_name = self.unit_var.get().strip() or "unit"
        width_min = float(self.width_min_var.get())
        width_max = float(self.width_max_var.get())
        width_step = float(self.width_step_var.get())
        exceed_width = float(self.exceed_width_var.get())
        bins = build_width_bins(width_min, width_max, width_step)
        threshold = int(float(self.threshold_var.get()))
        min_points = max(1, int(float(self.min_points_var.get())))
        return pixel_value, real_value, unit_name, bins, threshold, min_points, exceed_width

    def cache_key(self, image_path: Path) -> Tuple[str, str]:
        values = [
            self.pixel_value_var.get(),
            self.real_value_var.get(),
            self.unit_var.get(),
            self.width_min_var.get(),
            self.width_max_var.get(),
            self.width_step_var.get(),
            self.exceed_width_var.get(),
            self.threshold_var.get(),
            self.min_points_var.get(),
            str(self.auto_foreground_var.get()),
            str(self.crack_dark_var.get()),
        ]
        return (str(image_path), "|".join(values))

    def analyze_current(self) -> None:
        if not self.binary_images:
            self.status_var.set(f"未在 {self.binary_dir} 中找到二值化图。")
            return
        image_path = self.binary_images[self.index]
        try:
            pixel_value, real_value, unit_name, bins, threshold, min_points, exceed_width = self.current_parameters()
            key = self.cache_key(image_path)
            if key not in self.result_cache:
                self.result_cache[key] = analyze_binary_image(
                    image_path=image_path,
                    pixel_value=pixel_value,
                    real_value=real_value,
                    unit_name=unit_name,
                    bins=bins,
                    threshold=threshold,
                    auto_foreground=self.auto_foreground_var.get(),
                    crack_is_dark=self.crack_dark_var.get(),
                    min_skeleton_points=min_points,
                    exceed_width_threshold=exceed_width,
                )
            self.current_result = self.result_cache[key]
            self.selected_segment_id = None
            self.selected_bin_index = None
            self.update_counter()
            self.refresh_views()
            self.refresh_table()
            self.status_var.set(
                f"当前图：{len(self.current_result.segments)} 个分叉切段，"
                f"总长度 {self.current_result.total_length_unit:.4f} {unit_name}。"
            )
        except Exception as exc:
            messagebox.showerror("统计失败", str(exc))

    def refresh_views(self) -> None:
        if self.current_result is None:
            return
        try:
            pixel_value, real_value, _unit_name, bins, _threshold, _min_points, _exceed_width = self.current_parameters()
        except Exception:
            return
        binary_path = self.binary_images[self.index]
        original = read_display_image(self.original_map.get(binary_path.name), self.current_result.mask.shape)
        line_width = max(1, int(round(self.line_width_var.get())))
        overlay = draw_skeleton_pixels(
            original,
            self.current_result,
            bins,
            pixel_value,
            real_value,
            mode="blue",
            show_labels=self.show_labels_var.get(),
            alpha=self.overlay_alpha_var.get(),
            line_width=line_width,
            selected_segment_id=self.selected_segment_id,
            selected_bin_index=self.selected_bin_index,
        )
        binary_view = mask_to_display(self.current_result.mask, self.white_bg_var.get())
        colored = draw_skeleton_pixels(
            binary_view,
            self.current_result,
            bins,
            pixel_value,
            real_value,
            mode="bins",
            show_labels=self.show_labels_var.get(),
            alpha=1.0,
            line_width=line_width,
            selected_segment_id=self.selected_segment_id,
            selected_bin_index=self.selected_bin_index,
        )
        self.overlay_panel.title_var.set(f"叠加图：{self.current_result.name}")
        self.binary_panel.title_var.set(f"二值视图（分段着色）：{self.current_result.name}")
        self.overlay_panel.set_image(overlay)
        self.binary_panel.set_image(colored)
        self.refresh_legend()

    def refresh_table(self) -> None:
        for table in (self.segment_table, self.bin_table_widget, self.qc_table):
            for item in table.get_children():
                table.delete(item)
        if self.current_result is None:
            return
        unit_name = self.unit_var.get().strip() or "unit"
        self.segment_table.heading("length_unit", text=f"长度({unit_name})")
        self.segment_table.heading("width_unit", text=f"宽度({unit_name})")
        for seg in self.current_result.segments:
            self.segment_table.insert(
                "",
                "end",
                iid=f"seg-{seg.segment_id}",
                values=(
                    seg.segment_id,
                    seg.network_id,
                    f"{seg.length_px:.2f}",
                    f"{seg.width_px:.2f}",
                    f"{seg.length_unit:.4f}",
                    f"{seg.width_unit:.4f}",
                    seg.skeleton_points,
                ),
            )

        length_col = f"累计长度({unit_name})"
        for idx, row_data in enumerate(self.current_result.bin_table):
            self.bin_table_widget.insert(
                "",
                "end",
                iid=f"bin-{idx}",
                values=(
                    row_data["区间"],
                    f"{float(row_data[length_col]):.4f}",
                    f"{float(row_data['长度占比(%)']):.2f}",
                    f"{float(row_data['累计占比(%)']):.2f}",
                    f"{float(row_data['≥区间下限占比(%)']):.2f}",
                    row_data["骨架点数"],
                ),
            )

        for group_name, rows in [
            ("宽度统计", self.current_result.width_summary),
            ("质量控制", self.current_result.quality_metrics),
        ]:
            self.qc_table.insert("", "end", values=(f"[{group_name}]", ""))
            for key, value in rows.items():
                self.qc_table.insert("", "end", values=(key, format_value(value)))
        self.refresh_summary()

    def refresh_summary(self) -> None:
        if self.current_result is None:
            self.summary_var.set("")
            return
        unit_name = self.unit_var.get().strip() or "unit"
        summary = self.current_result.width_summary
        quality = self.current_result.quality_metrics
        mean_col = f"长度加权平均宽度({unit_name})"
        max_col = f"最大宽度({unit_name})"
        threshold_col = f"超阈值宽度阈值({unit_name})"
        self.summary_var.set(
            "\n".join(
                [
                    f"总长度：{self.current_result.total_length_unit:.4f} {unit_name}",
                    f"平均宽度：{float(summary.get(mean_col, 0.0)):.4f} {unit_name}",
                    f"最大宽度：{float(summary.get(max_col, 0.0)):.4f} {unit_name}",
                    f"主要区间：{main_width_bin_label(self.current_result.bin_table, unit_name)}",
                    f"≥{float(summary.get(threshold_col, 0.0)):g} {unit_name}：{float(summary.get('超阈值长度占比(%)', 0.0)):.2f}%",
                    f"质检：{quality.get('质检提示', '-')}",
                ]
            )
        )

    def refresh_legend(self) -> None:
        if not hasattr(self, "legend_canvas"):
            return
        self.legend_canvas.delete("all")
        if self.current_result is None:
            return
        try:
            _pixel_value, _real_value, unit_name, bins, _threshold, _min_points, _exceed_width = self.current_parameters()
        except Exception:
            return
        colors = colors_for_bins(len(bins))
        width = max(self.legend_canvas.winfo_width(), 200)
        item_width = 128
        x = 10
        y = 8
        for idx, ((low, high), color) in enumerate(zip(bins, colors)):
            if x + item_width > width:
                x = 10
                y += 24
            color_hex = f"#{color[0]:02x}{color[1]:02x}{color[2]:02x}"
            outline = "#ff2814" if self.selected_bin_index == idx else "#666"
            self.legend_canvas.create_rectangle(x, y, x + 18, y + 14, fill=color_hex, outline=outline, width=2)
            self.legend_canvas.create_text(x + 24, y + 7, text=f"{low:g}-{high:g} {unit_name}", anchor="w", fill="#222")
            x += item_width

    def on_segment_select(self, _event: tk.Event) -> None:
        selected = self.segment_table.selection()
        if not selected:
            return
        item_id = selected[0]
        if not item_id.startswith("seg-"):
            return
        self.selected_segment_id = int(item_id.split("-", 1)[1])
        self.selected_bin_index = None
        self.bin_table_widget.selection_remove(self.bin_table_widget.selection())
        self.refresh_views()

    def on_bin_select(self, _event: tk.Event) -> None:
        selected = self.bin_table_widget.selection()
        if not selected:
            return
        item_id = selected[0]
        if not item_id.startswith("bin-"):
            return
        self.selected_bin_index = int(item_id.split("-", 1)[1])
        self.selected_segment_id = None
        self.segment_table.selection_remove(self.segment_table.selection())
        self.refresh_views()

    def prev_image(self) -> None:
        if not self.binary_images:
            return
        self.index = (self.index - 1) % len(self.binary_images)
        self.analyze_current()

    def next_image(self) -> None:
        if not self.binary_images:
            return
        self.index = (self.index + 1) % len(self.binary_images)
        self.analyze_current()

    def show_quality_preview(self) -> None:
        if not self.binary_images:
            messagebox.showwarning("无法预览", "没有找到二值化图。")
            return
        try:
            pixel_value, real_value, unit_name, bins, threshold, min_points, exceed_width = self.current_parameters()
        except Exception as exc:
            messagebox.showerror("参数错误", str(exc))
            return

        top = tk.Toplevel(self)
        top.title("批量质检预览")
        top.geometry("1180x760")
        container = ttk.Frame(top)
        container.pack(fill="both", expand=True)
        canvas = tk.Canvas(container, bg="#f4f4f4", highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = ttk.Frame(canvas)
        window_id = canvas.create_window((0, 0), window=inner, anchor="nw")
        self._quality_preview_images = []

        def update_scrollregion(_event: tk.Event) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))

        def update_width(event: tk.Event) -> None:
            canvas.itemconfigure(window_id, width=event.width)

        inner.bind("<Configure>", update_scrollregion)
        canvas.bind("<Configure>", update_width)

        columns = 3
        line_width = max(1, int(round(self.line_width_var.get())))
        for idx, binary_path in enumerate(self.binary_images):
            result = analyze_binary_image(
                image_path=binary_path,
                pixel_value=pixel_value,
                real_value=real_value,
                unit_name=unit_name,
                bins=bins,
                threshold=threshold,
                auto_foreground=self.auto_foreground_var.get(),
                crack_is_dark=self.crack_dark_var.get(),
                min_skeleton_points=min_points,
                exceed_width_threshold=exceed_width,
            )
            original = read_display_image(self.original_map.get(binary_path.name), result.mask.shape)
            overlay = draw_skeleton_pixels(
                original,
                result,
                bins,
                pixel_value,
                real_value,
                mode="blue",
                show_labels=False,
                alpha=self.overlay_alpha_var.get(),
                line_width=line_width,
            )
            overlay.thumbnail((330, 190), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(overlay)
            self._quality_preview_images.append(photo)

            row = idx // columns
            col = idx % columns
            cell = ttk.Frame(inner, padding=8)
            cell.grid(row=row, column=col, sticky="nsew", padx=6, pady=6)
            ttk.Label(cell, image=photo).pack()
            mean_col = f"长度加权平均宽度({unit_name})"
            max_col = f"最大宽度({unit_name})"
            text = (
                f"{result.name}\n"
                f"总长度：{result.total_length_unit:.4f} {unit_name}\n"
                f"平均宽度：{float(result.width_summary.get(mean_col, 0.0)):.4f} {unit_name}；"
                f"最大：{float(result.width_summary.get(max_col, 0.0)):.4f} {unit_name}\n"
                f"质检：{result.quality_metrics.get('质检提示', '-')}"
            )
            ttk.Label(cell, text=text, justify="left", wraplength=330).pack(fill="x", pady=(6, 0))
            self.status_var.set(f"已生成质检预览 {idx + 1}/{len(self.binary_images)}")
            self.update_idletasks()

        for col in range(columns):
            inner.columnconfigure(col, weight=1)
        self.status_var.set("批量质检预览已生成。")

    def export_all(self) -> None:
        if not self.binary_images:
            messagebox.showwarning("无法导出", "没有找到二值化图。")
            return
        try:
            pixel_value, real_value, unit_name, bins, threshold, min_points, exceed_width = self.current_parameters()
        except Exception as exc:
            messagebox.showerror("参数错误", str(exc))
            return

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = self.data_dir / f"裂纹统计输出_{stamp}"
        preview_dir = out_dir / "预览图"
        preview_dir.mkdir(parents=True, exist_ok=True)
        excel_path = out_dir / "crack_width_statistics.xlsx"

        all_segment_frames = []
        all_bin_frames = []
        summary_rows = []
        global_width_values = []
        global_weights = []
        preview_tiles: List[Tuple[str, Image.Image, str]] = []

        try:
            workbook = Workbook()
            workbook.remove(workbook.active)
            segment_cols = segment_columns(unit_name)
            line_width = max(1, int(round(self.line_width_var.get())))

            for idx, binary_path in enumerate(self.binary_images, start=1):
                result = analyze_binary_image(
                    image_path=binary_path,
                    pixel_value=pixel_value,
                    real_value=real_value,
                    unit_name=unit_name,
                    bins=bins,
                    threshold=threshold,
                    auto_foreground=self.auto_foreground_var.get(),
                    crack_is_dark=self.crack_dark_var.get(),
                    min_skeleton_points=min_points,
                    exceed_width_threshold=exceed_width,
                )
                seg_rows = segments_to_rows(result.segments, unit_name)
                bin_rows = list(result.bin_table)

                sheet = workbook.create_sheet(unique_sheet_name(workbook, result.name))
                next_row = write_table(sheet, seg_rows, 1, segment_cols)
                next_row = write_table(sheet, bin_rows, next_row + 2)
                next_row = write_table(sheet, dict_to_rows(result.width_summary), next_row + 2, ["指标", "数值"])
                write_table(sheet, dict_to_rows(result.quality_metrics), next_row + 2, ["指标", "数值"])

                for row_data in seg_rows:
                    named_row = {"图片名": result.name}
                    named_row.update(row_data)
                    all_segment_frames.append(named_row)

                for row_data in bin_rows:
                    named_row = {"图片名": result.name}
                    named_row.update(row_data)
                    all_bin_frames.append(named_row)

                summary_row = {
                    "图片名": result.name,
                    "裂纹切段数": len(result.segments),
                    "主要宽度区间": main_width_bin_label(result.bin_table, unit_name),
                }
                summary_row.update(result.width_summary)
                summary_row.update(result.quality_metrics)
                summary_rows.append(summary_row)

                global_width_values.append(result.widths_px[result.skeleton] * unit_factor(pixel_value, real_value))
                global_weights.append(result.pixel_length_share[result.skeleton])

                original = read_display_image(self.original_map.get(binary_path.name), result.mask.shape)
                overlay = draw_skeleton_pixels(
                    original,
                    result,
                    bins,
                    pixel_value,
                    real_value,
                    mode="blue",
                    show_labels=self.show_labels_var.get(),
                    alpha=self.overlay_alpha_var.get(),
                    line_width=line_width,
                )
                overlay.save(preview_dir / f"{result.name}_overlay.png")
                colored = draw_skeleton_pixels(
                    mask_to_display(result.mask, self.white_bg_var.get()),
                    result,
                    bins,
                    pixel_value,
                    real_value,
                    mode="bins",
                    show_labels=self.show_labels_var.get(),
                    alpha=1.0,
                    line_width=line_width,
                )
                colored.save(preview_dir / f"{result.name}_width_bins.png")

                tile = overlay.copy()
                tile.thumbnail((320, 190), Image.Resampling.LANCZOS)
                mean_col = f"长度加权平均宽度({unit_name})"
                tile_text = (
                    f"{result.name}\n"
                    f"L={result.total_length_unit:.4f} {unit_name}; "
                    f"Wmean={float(result.width_summary.get(mean_col, 0.0)):.4f} {unit_name}\n"
                    f"QC: {result.quality_metrics.get('质检提示', '-')}"
                )
                preview_tiles.append((result.name, tile, tile_text))
                self.status_var.set(f"正在导出 {idx}/{len(self.binary_images)}：{result.name}")
                self.update_idletasks()

            if all_segment_frames:
                write_table(workbook.create_sheet("总汇总表"), all_segment_frames, 1, ["图片名"] + segment_cols)
            if all_bin_frames:
                write_table(workbook.create_sheet("各图宽度区间统计"), all_bin_frames, 1)
                lower_col = f"区间下限({unit_name})"
                upper_col = f"区间上限({unit_name})"
                length_col = f"累计长度({unit_name})"
                above_col = f"≥区间下限长度({unit_name})"
                grouped_rows: Dict[Tuple[object, object], Dict[str, object]] = {}
                for row_data in all_bin_frames:
                    key = (row_data[lower_col], row_data[upper_col])
                    if key not in grouped_rows:
                        grouped_rows[key] = {
                            "区间": row_data["区间"],
                            lower_col: row_data[lower_col],
                            upper_col: row_data[upper_col],
                            "累计长度(px)": 0.0,
                            length_col: 0.0,
                            above_col: 0.0,
                            "骨架点数": 0,
                        }
                    grouped_rows[key]["累计长度(px)"] += float(row_data["累计长度(px)"])
                    grouped_rows[key][length_col] += float(row_data[length_col])
                    grouped_rows[key][above_col] += float(row_data[above_col])
                    grouped_rows[key]["骨架点数"] += int(row_data["骨架点数"])
                grouped_list = sorted(grouped_rows.values(), key=lambda row: float(row[lower_col]))
                total_global_length_px = sum(float(row.get("总长度(px)", 0.0)) for row in summary_rows)
                total_global_length_unit = sum(float(row.get(f"总长度({unit_name})", 0.0)) for row in summary_rows)
                cumulative_px = 0.0
                for row in grouped_list:
                    cumulative_px += float(row["累计长度(px)"])
                    row["长度占比(%)"] = percent(float(row["累计长度(px)"]), total_global_length_px)
                    row["累计占比(%)"] = percent(cumulative_px, total_global_length_px)
                    row["≥区间下限占比(%)"] = percent(float(row[above_col]), total_global_length_unit)
                write_table(workbook.create_sheet("宽度区间总统计"), grouped_list, 1)

            if global_width_values:
                values = np.concatenate(global_width_values)
                weights = np.concatenate(global_weights)
                total_weight = float(weights.sum())
                if total_weight > 0:
                    mean = float(np.average(values, weights=weights))
                    variance = float(np.average((values - mean) ** 2, weights=weights))
                    exceed_length = float(weights[values >= exceed_width].sum())
                    global_summary = {
                        f"总长度({unit_name})": total_weight * unit_factor(pixel_value, real_value),
                        f"长度加权平均宽度({unit_name})": mean,
                        f"中位宽度({unit_name})": weighted_percentile(values, weights, 50.0),
                        f"P10宽度({unit_name})": weighted_percentile(values, weights, 10.0),
                        f"P90宽度({unit_name})": weighted_percentile(values, weights, 90.0),
                        f"最大宽度({unit_name})": float(values.max()),
                        f"宽度标准差({unit_name})": math.sqrt(max(variance, 0.0)),
                        f"超阈值宽度阈值({unit_name})": exceed_width,
                        "超阈值长度占比(%)": percent(exceed_length, total_weight),
                    }
                    write_table(workbook.create_sheet("全局宽度统计"), dict_to_rows(global_summary), 1, ["指标", "数值"])
            write_table(workbook.create_sheet("图片总览"), summary_rows, 1)

            if preview_tiles:
                tile_w, tile_h = 380, 280
                columns = 3
                rows = math.ceil(len(preview_tiles) / columns)
                contact = Image.new("RGB", (tile_w * columns, tile_h * rows), "white")
                draw = ImageDraw.Draw(contact)
                for idx, (_name, tile, text) in enumerate(preview_tiles):
                    x = (idx % columns) * tile_w
                    y = (idx // columns) * tile_h
                    contact.paste(tile.convert("RGB"), (x + 12, y + 12))
                    draw.text((x + 12, y + 210), text, fill=(20, 20, 20))
                    draw.rectangle((x, y, x + tile_w - 1, y + tile_h - 1), outline=(210, 210, 210))
                contact.save(out_dir / "batch_quality_preview.png")
            workbook.save(excel_path)

            self.status_var.set(f"导出完成：{excel_path}")
            messagebox.showinfo("导出完成", f"已导出：\n{excel_path}\n\n预览图目录：\n{preview_dir}\n\n批量质检图：\n{out_dir / 'batch_quality_preview.png'}")
        except Exception as exc:
            messagebox.showerror("导出失败", str(exc))

    @staticmethod
    def safe_sheet_name(name: str) -> str:
        cleaned = re.sub(r"[\[\]:*?/\\]", "_", name)
        return cleaned[:31] or "Sheet"


def main() -> None:
    if "--self-test" in sys.argv:
        args = sys.argv[1:]
        pos = args.index("--self-test")
        data_dir = Path(args[pos + 1]) if pos + 1 < len(args) else DEFAULT_DATA_DIR
        output_dir = Path(args[pos + 2]) if pos + 2 < len(args) else Path.cwd()
        raise SystemExit(run_self_test(data_dir, output_dir))
    app = CrackStatsApp()
    app.mainloop()


if __name__ == "__main__":
    main()
